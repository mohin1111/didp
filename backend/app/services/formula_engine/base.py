"""
Base Excel formula functions - core operations like SUM, AVERAGE, VLOOKUP, etc.
"""
import tempfile
import os
from typing import List, Any

import pandas as pd
import numpy as np


class BaseFormulaMixin:
    """Base mixin with core Excel formula functions."""

    def __init__(self):
        self._pycel_available = False
        try:
            from pycel import ExcelCompiler
            from openpyxl import Workbook
            self._pycel_available = True
        except ImportError:
            pass

    def evaluate(self, formula: str, data: pd.DataFrame, row: int = 0) -> Any:
        """Evaluate an Excel formula against DataFrame data."""
        if not self._pycel_available:
            raise RuntimeError("pycel not available. Install with: pip install pycel")

        from pycel import ExcelCompiler
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_idx, col_name in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row_data in enumerate(data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                try:
                    ws.cell(row=row_idx, column=col_idx, value=float(value) if value != '' else None)
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        formula_row = len(data) + 3
        ws.cell(row=formula_row, column=1, value=formula)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
            wb.save(temp_path)

        try:
            excel = ExcelCompiler(filename=temp_path)
            result = excel.evaluate(f"Data!A{formula_row}")
            return result
        finally:
            os.unlink(temp_path)

    def evaluate_column(self, formula: str, data: pd.DataFrame) -> List[Any]:
        """Evaluate a formula for each row, creating a new column."""
        if not self._pycel_available:
            raise RuntimeError("pycel not available")

        from pycel import ExcelCompiler
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        for col_idx, col_name in enumerate(data.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, row_data in enumerate(data.values, 2):
            for col_idx, value in enumerate(row_data, 1):
                try:
                    ws.cell(row=row_idx, column=col_idx, value=float(value) if value != '' else None)
                except (ValueError, TypeError):
                    ws.cell(row=row_idx, column=col_idx, value=value)

        formula_col = len(data.columns) + 1
        for row_idx in range(2, len(data) + 2):
            adjusted_formula = formula.replace('2', str(row_idx))
            ws.cell(row=row_idx, column=formula_col, value=adjusted_formula)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name
            wb.save(temp_path)

        try:
            excel = ExcelCompiler(filename=temp_path)
            col_letter = self._col_to_letter(formula_col)

            results = []
            for row_idx in range(2, len(data) + 2):
                try:
                    result = excel.evaluate(f"Data!{col_letter}{row_idx}")
                    results.append(result)
                except Exception:
                    results.append(None)

            return results
        finally:
            os.unlink(temp_path)

    def apply_formula(self, data: pd.DataFrame, column_name: str, formula: str) -> pd.DataFrame:
        """Add a new column to DataFrame by applying an Excel formula."""
        results = self.evaluate_column(formula, data)
        data = data.copy()
        data[column_name] = results
        return data

    def sum_range(self, data: pd.DataFrame, column: str) -> float:
        """SUM function for a column."""
        return pd.to_numeric(data[column], errors='coerce').sum()

    def average(self, data: pd.DataFrame, column: str) -> float:
        """AVERAGE function for a column."""
        return pd.to_numeric(data[column], errors='coerce').mean()

    def count(self, data: pd.DataFrame, column: str) -> int:
        """COUNT function (numeric values only)."""
        return pd.to_numeric(data[column], errors='coerce').notna().sum()

    def counta(self, data: pd.DataFrame, column: str) -> int:
        """COUNTA function (all non-empty values)."""
        return data[column].notna().sum()

    def max_val(self, data: pd.DataFrame, column: str) -> float:
        """MAX function for a column."""
        return pd.to_numeric(data[column], errors='coerce').max()

    def min_val(self, data: pd.DataFrame, column: str) -> float:
        """MIN function for a column."""
        return pd.to_numeric(data[column], errors='coerce').min()

    def sumif(self, data: pd.DataFrame, criteria_col: str, criteria: Any, sum_col: str) -> float:
        """SUMIF function."""
        mask = data[criteria_col] == criteria
        return pd.to_numeric(data.loc[mask, sum_col], errors='coerce').sum()

    def countif(self, data: pd.DataFrame, column: str, criteria: Any) -> int:
        """COUNTIF function."""
        return (data[column] == criteria).sum()

    def vlookup(self, lookup_value: Any, data: pd.DataFrame, col_index: int, exact_match: bool = True) -> Any:
        """VLOOKUP function."""
        first_col = data.columns[0]
        if exact_match:
            mask = data[first_col] == lookup_value
            matches = data.loc[mask]
            if len(matches) > 0:
                return matches.iloc[0, col_index - 1]
            return None
        else:
            numeric_data = pd.to_numeric(data[first_col], errors='coerce')
            try:
                lookup_num = float(lookup_value)
                valid_mask = numeric_data <= lookup_num
                if valid_mask.any():
                    idx = numeric_data[valid_mask].idxmax()
                    return data.iloc[idx, col_index - 1]
            except (ValueError, TypeError):
                pass
            return None

    def index_match(self, data: pd.DataFrame, match_col: str, match_value: Any, return_col: str) -> Any:
        """INDEX/MATCH combination."""
        mask = data[match_col] == match_value
        matches = data.loc[mask, return_col]
        if len(matches) > 0:
            return matches.iloc[0]
        return None

    def _col_to_letter(self, col_num: int) -> str:
        """Convert column number to Excel letter (1=A, 2=B, etc.)."""
        result = ""
        while col_num > 0:
            col_num, remainder = divmod(col_num - 1, 26)
            result = chr(65 + remainder) + result
        return result

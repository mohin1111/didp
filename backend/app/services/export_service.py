from io import BytesIO
from typing import List, Optional
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models import Table, TableColumn, TableRow, MatchResult


class ExportService:
    """Service for exporting data to Excel/CSV"""

    def __init__(self, db: Session):
        self.db = db

    def export_tables_to_excel(
        self,
        table_keys: List[str],
        include_headers: bool = True
    ) -> BytesIO:
        """Export multiple tables to Excel workbook (each as sheet)"""
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        for table_key in table_keys:
            table = self.db.query(Table).filter(Table.key == table_key).first()
            if not table:
                continue

            # Create worksheet
            ws = wb.create_sheet(title=table.name[:31])  # Excel sheet name limit

            # Get columns and data
            columns = [c.name for c in sorted(table.columns, key=lambda x: x.index)]
            rows = [r.data for r in sorted(table.data_rows, key=lambda x: x.row_index)]

            # Write headers
            if include_headers:
                for col_idx, col_name in enumerate(columns, 1):
                    ws.cell(row=1, column=col_idx, value=col_name)

            # Write data
            start_row = 2 if include_headers else 1
            for row_idx, row_data in enumerate(rows, start_row):
                for col_idx, cell_value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)

            # Auto-size columns
            for col_idx, col_name in enumerate(columns, 1):
                max_length = len(str(col_name))
                for row in rows[:100]:  # Sample first 100 rows
                    if col_idx - 1 < len(row):
                        max_length = max(max_length, len(str(row[col_idx - 1] or '')))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        # If no sheets were created, add an empty one
        if len(wb.sheetnames) == 0:
            wb.create_sheet(title="Empty")

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_table_to_csv(self, table_key: str) -> BytesIO:
        """Export single table to CSV"""
        import csv

        table = self.db.query(Table).filter(Table.key == table_key).first()
        if not table:
            raise ValueError(f"Table '{table_key}' not found")

        columns = [c.name for c in sorted(table.columns, key=lambda x: x.index)]
        rows = [r.data for r in sorted(table.data_rows, key=lambda x: x.row_index)]

        output = BytesIO()
        # Use TextIOWrapper for CSV writer
        import io
        text_output = io.StringIO()

        writer = csv.writer(text_output)
        writer.writerow(columns)
        writer.writerows(rows)

        output.write(text_output.getvalue().encode('utf-8'))
        output.seek(0)
        return output

    def export_match_results(
        self,
        result_id: int,
        include_matched: bool = True,
        include_unmatched: bool = True
    ) -> BytesIO:
        """Export match results to Excel with multiple sheets"""
        result = self.db.query(MatchResult).filter(MatchResult.id == result_id).first()
        if not result:
            raise ValueError(f"Match result {result_id} not found")

        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Summary sheet
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary.append(["Match Result Summary"])
        ws_summary.append([""])
        ws_summary.append(["Config ID", result.config_id])
        ws_summary.append(["Matched Count", result.matched_count])
        ws_summary.append(["Unmatched Source", result.unmatched_source_count])
        ws_summary.append(["Unmatched Target", result.unmatched_target_count])
        ws_summary.append(["Created At", str(result.created_at)])

        # Matched pairs sheet
        if include_matched and result.matched_pairs:
            ws_matched = wb.create_sheet(title="Matched Pairs")

            # Get max columns from data
            if result.matched_pairs:
                sample = result.matched_pairs[0]
                source_cols = len(sample.get("source_row", []))
                target_cols = len(sample.get("target_row", []))

                # Headers
                headers = ["Source Row #"] + [f"Source Col {i+1}" for i in range(source_cols)]
                headers += ["Target Row #"] + [f"Target Col {i+1}" for i in range(target_cols)]
                ws_matched.append(headers)

                # Data
                for pair in result.matched_pairs:
                    row = [pair["source_row_index"]] + pair["source_row"]
                    row += [pair["target_row_index"]] + pair["target_row"]
                    ws_matched.append(row)

        # Unmatched source sheet
        if include_unmatched and result.unmatched_source:
            ws_unmatched_src = wb.create_sheet(title="Unmatched Source")

            if result.unmatched_source:
                sample = result.unmatched_source[0]
                cols = len(sample.get("row", []))

                headers = ["Row #"] + [f"Column {i+1}" for i in range(cols)]
                ws_unmatched_src.append(headers)

                for item in result.unmatched_source:
                    row = [item["row_index"]] + item["row"]
                    ws_unmatched_src.append(row)

        # Unmatched target sheet
        if include_unmatched and result.unmatched_target:
            ws_unmatched_tgt = wb.create_sheet(title="Unmatched Target")

            if result.unmatched_target:
                sample = result.unmatched_target[0]
                cols = len(sample.get("row", []))

                headers = ["Row #"] + [f"Column {i+1}" for i in range(cols)]
                ws_unmatched_tgt.append(headers)

                for item in result.unmatched_target:
                    row = [item["row_index"]] + item["row"]
                    ws_unmatched_tgt.append(row)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_sql_results(
        self,
        columns: List[str],
        data: List[List[str]],
        filename: str = "SQL_Results"
    ) -> BytesIO:
        """Export SQL query results to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "SQL Results"

        # Write headers
        for col_idx, col_name in enumerate(columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        # Auto-size columns
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row in data[:100]:  # Sample first 100 rows
                if col_idx - 1 < len(row):
                    max_length = max(max_length, len(str(row[col_idx - 1] or '')))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def export_comparison(
        self,
        rows: List[dict],
        table_names: dict
    ) -> BytesIO:
        """Export comparison rows to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison"

        if not rows:
            ws.append(["No data to compare"])
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output

        # Calculate max columns
        max_cols = max(len(row.get("data", [])) for row in rows)
        headers = ["Source Table", "Row #"] + [f"Column {i + 1}" for i in range(max_cols)]

        # Write headers
        ws.append(headers)

        # Write data
        for row in rows:
            table_key = row.get("table_key", "")
            table_name = table_names.get(table_key, table_key)
            row_index = row.get("row_index", 0) + 1
            data = row.get("data", [])

            ws.append([table_name, row_index] + data)

        # Auto-size columns
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
